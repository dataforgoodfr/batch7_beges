import numpy as np
import pandas as pd
import io
from tempfile import NamedTemporaryFile
import openpyxl
import xlrd
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import load_workbook
from utils.organization_chart import oc
from utils.chorus_dt_handler import ch
from utils.odrive_handler import ov
from utils.osfi_handler import oh

EXCEL_TEMPLATE_PATH = "/data/templates/beges_template.xlsx"


class DataExport:
    def __init__(self, selected_entity):
        self.service = oc.get_entity_by_id(selected_entity)
        self.load_data()

    def load_data(self):
        self.chorus_dt_df = ch.get_structure_data(self.service.code_chorus)
        self.odrive_df = ov.get_structure_data(self.service.code_odrive)
        self.osfi_df = oh.get_structure_data(self.service.code_osfi)
        pass

    def get_file_as_bytes(self):
        """Function returns Excel data as bytes array. It avoids the need to create a file in memory.
            See https://xlsxwriter.readthedocs.io/working_with_pandas.html#additional-pandas-and-excel-information"""
        output = io.BytesIO()
        with pd.ExcelWriter(
            output, engine="xlsxwriter"
        ) as writer:  # https://community.plotly.com/t/allow-users-to-dowload-an-excel-in-a-click/9410
            self.chorus_dt_df.to_excel(writer, sheet_name="data_chorus_dt", index_label="index")
            self.odrive_df.to_excel(writer, sheet_name="data_odrive", index_label="index")
            self.osfi_df.to_excel(writer, sheet_name="data_osfi", index_label="index")
        writer.save()
        xlsx_data = output.getvalue()
        output.seek(0)
        return output

    def get_file_as_bytes_openpyxl(self):
        """Function returns Excel data as bytes array. It avoids the need to create a file in memory.
            https://openpyxl.readthedocs.io/en/stable/tutorial.html"""
        import time

        start = time.time()
        wb = load_workbook(EXCEL_TEMPLATE_PATH)
        print("INFO: Template file loaded. Time: {}".format(time.time() - start))
        with NamedTemporaryFile() as tmp:
            ws = wb.create_sheet(title="data_chorus_dt")
            for r in dataframe_to_rows(self.chorus_dt_df, header=True):
                ws.append(r)
            print("INFO: Workbook filled. Time: {}".format(time.time() - start))
            wb.save(tmp.name)
            print("INFO: Workbook saved. Time: {}".format(time.time() - start))
            tmp.seek(0)
            bytes = tmp.read()
        print("INFO: File created. Time: {}".format(time.time() - start))
        return io.BytesIO(bytes)
